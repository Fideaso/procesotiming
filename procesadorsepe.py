import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import threading
from tkinter.ttk import Progressbar, Style, Button

def eliminar_filtros(sheet):
    if sheet.auto_filter:
        sheet.auto_filter.ref = None

def buscar_encabezado_id_guia(sheet):
    for cell in sheet[2]:
        if cell.value == "ID GUIA":
            return cell.column
    
    return None

def eliminar_filas_vacias_y_anterior_id_guia(sheet, columna_id_guia, progress_bar):
    registros_eliminados = 0
    
    if columna_id_guia > 1:
        sheet.delete_rows(columna_id_guia - 1)
        registros_eliminados += 1
    
    total_filas = sheet.max_row - columna_id_guia
    progress_value = 0
    
    for row in range(sheet.max_row, columna_id_guia, -1):
        if sheet.cell(row=row, column=columna_id_guia).value is None:
            fila_vacia = True
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=row, column=col).value is not None:
                    fila_vacia = False
                    break
            
            if not fila_vacia:
                sheet.delete_rows(row)
                registros_eliminados += 1
        
        progress_value += 1
        progress_bar["value"] = (progress_value / total_filas) * 100
        progress_bar.update()
    
    messagebox.showinfo("Proceso completado", f"Se han eliminado {registros_eliminados} filas según las condiciones especificadas.")
    return registros_eliminados

def eliminar_encabezado_id_guia_y_filtros(archivo_entrada, archivo_salida, progress_bar):
    wb = openpyxl.load_workbook(archivo_entrada)
    sheet = wb.active
    
    eliminar_filtros(sheet)
    columna_id_guia = buscar_encabezado_id_guia(sheet)
    
    if columna_id_guia is not None:
        registros_eliminados = eliminar_filas_vacias_y_anterior_id_guia(sheet, columna_id_guia, progress_bar)
        wb.save(archivo_salida)
        progress_bar.destroy()

def procesar_archivo(archivo_entrada, archivo_salida, progress_bar):
    eliminar_encabezado_id_guia_y_filtros(archivo_entrada, archivo_salida, progress_bar)
    root.after(0, lambda: root.focus_force())

def seleccionar_archivo():
    archivo_entrada = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    
    if archivo_entrada:
        if not archivo_entrada.endswith('.xlsx'):
            messagebox.showerror("Error", "El archivo seleccionado no es un archivo .xlsx válido.")
            return
        
        archivo_salida = archivo_entrada.replace('.xlsx', '_procesado.xlsx')
        
        progress_bar = Progressbar(root, orient=tk.HORIZONTAL, mode='determinate')
        progress_bar.pack(fill=tk.X)
        
        thread = threading.Thread(target=procesar_archivo, args=(archivo_entrada, archivo_salida, progress_bar))
        thread.start()

root = tk.Tk()
root.title("Eliminar filas vacías y fila anterior a 'ID GUIA' y filtros de archivo Excel")

style = Style()
style.theme_use("clam")
style.configure("TButton", padding=10, relief="flat", background="#4CAF50", foreground="white", font=("Helvetica", 12))
style.map("TButton", background=[("active", "#45a049")])

boton_seleccionar = Button(root, text="Seleccionar archivo", command=seleccionar_archivo, style="TButton")
boton_seleccionar.pack(pady=20)

root.mainloop()
