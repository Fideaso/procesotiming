import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.ttk import Progressbar, Style, Button
import openpyxl
import threading
import re
import datetime
from openpyxl.utils import get_column_letter

def procesar_archivo():
    # Abre el diálogo para seleccionar el archivo xlsx
    ruta_archivo = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    
    if ruta_archivo:
        try:
            # Carga el libro de trabajo
            libro = openpyxl.load_workbook(filename=ruta_archivo)
            hoja = libro.active
            
            eliminar_encabezado_id_guia_y_filtros(hoja)
            procesar_fechas(hoja)
            
            # Guarda los cambios
            libro.save(ruta_archivo)
            
            mensaje_estado.config(text="Procesamiento completado.")
        except Exception as e:
            messagebox.showerror("Error", str(e))
    else:
        messagebox.showwarning("Advertencia", "No se seleccionó ningún archivo.")

def eliminar_encabezado_id_guia_y_filtros(hoja):
    eliminar_filtros(hoja)
    columna_id_guia = buscar_encabezado_id_guia(hoja)
    
    if columna_id_guia is not None:
        eliminar_filas_vacias_y_anterior_id_guia(hoja, columna_id_guia)

def eliminar_filtros(sheet):
    if sheet.auto_filter:
        sheet.auto_filter.ref = None

def buscar_encabezado_id_guia(sheet):
    for cell in sheet[2]:
        if cell.value == "ID GUIA":
            return cell.column
    return None

def eliminar_filas_vacias_y_anterior_id_guia(sheet, columna_id_guia):
    registros_eliminados = 0
    
    if columna_id_guia > 1:
        sheet.delete_rows(columna_id_guia - 1)
        registros_eliminados += 1
    
    total_filas = sheet.max_row - columna_id_guia
    progress_value = 0
    
    for row in range(sheet.max_row, columna_id_guia, -1):
        if sheet.cell(row=row, column=columna_id_guia).value is None:
            fila_vacia = True
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=row, column=col).value is not None:
                    fila_vacia = False
                    break
            
            if not fila_vacia:
                sheet.delete_rows(row)
                registros_eliminados += 1
        
        progress_value += 1
    
    messagebox.showinfo("Proceso completado", f"Se han eliminado {registros_eliminados} filas según las condiciones especificadas.")

def procesar_fechas(hoja):
    # Itera sobre las filas y columnas
    for row in hoja.iter_rows(min_row=2, max_row=hoja.max_row, min_col=28, max_col=29):
        for cell in row:
            if isinstance(cell.value, datetime.datetime):
                cell.value = cell.value.strftime('%d/%m/%Y')
                cell.number_format = '@'  # Formato de texto

def convertir_a_numero(archivo):
    wb = openpyxl.load_workbook(filename=archivo)
    ws = wb.active

    # Obtener la columna 'AA'
    columna = ws['AA']

    # Iterar sobre las celdas de la columna y convertirlas a números si es posible
    for celda in columna:
        if isinstance(celda.value, str) and celda.value.replace('.', '', 1).isdigit():
            celda.value = float(celda.value)

    # Guardar el archivo modificado
    wb.save(archivo)

# Configuración de la interfaz gráfica
root = tk.Tk()
root.title("Procesador de Archivos Excel")
root.geometry("400x150")

style = Style()
style.theme_use("clam")
style.configure("TButton", padding=10, relief="flat", background="#4CAF50", foreground="white", font=("Helvetica", 12))
style.map("TButton", background=[("active", "#45a049")])

btn_seleccionar = Button(root, text="Seleccionar Archivo", command=procesar_archivo, style="TButton")
btn_seleccionar.pack(pady=20)

mensaje_estado = tk.Label(root, text="")
mensaje_estado.pack()

# Ejecutar el bucle principal de la aplicación
root.mainloop()
